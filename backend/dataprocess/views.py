import os
import re
from django.contrib import auth
from django.shortcuts import render,get_object_or_404
from account.models import User
from django.http import HttpResponse

from dataprocess.models import CollectData
from crawler.models import *
from config.models import PlatformTargetItem, CollectTargetItem, Schedule
from config.serializers import PlatformTargetItemSerializer, CollectTargetItemSerializer, ScheduleSerializer
from dataprocess.functions import export_datareport, import_datareport, import_collects, import_authinfo
from dataprocess.pagination import ViewPaginatorMixin
from crawler.views import get_task_result, parse_logfile_for_error

from .serializers import *
from .models import *
from django.http.response import JsonResponse
from rest_framework.parsers import JSONParser
from rest_framework import status
from django.views.decorators.csrf import csrf_exempt
from utils.api import APIView, validate_serializer
from utils.shortcuts import get_env

import datetime
from datetime import timedelta
import openpyxl
from openpyxl.writer.excel import save_virtual_workbook
from django.http import FileResponse
from django.core.files.storage import FileSystemStorage
import logging
from django_celery_beat.models import PeriodicTask

formatter = logging.Formatter('[%(asctime)s] - [%(levelname)s] - [%(name)s:%(lineno)d]  - %(message)s', '%Y-%m-%d %H:%M:%S')
serverlogger = logging.getLogger(__name__)
userlogger = logging.getLogger("HTTP-Method")

production_env = get_env("YG_ENV", "dev") == "production"
if production_env:
    LOG_PATH = "/data/log/user"
else:
    LOG_PATH = "./data/log/user"

trfh = logging.handlers.TimedRotatingFileHandler(
    filename = os.path.join(LOG_PATH, f"{datetime.datetime.today().strftime('%Y-%m-%d')}.log"),
    when = "midnight",
    interval=1,
    encoding="utf-8",
)
trfh.setFormatter(formatter)
trfh.setLevel(logging.INFO)
userlogger.addHandler(trfh)
userlogger.setLevel(logging.DEBUG)

# 정의 : logincheck
# 목적 : 웹사이트의 쿠키를 보고 로그인 유무 판별
# 멤버함수 : 
# 개발자 : 김민희, minheekim3@naver.com
# 최종수정일 : 2022-02-22
def logincheck(request):
    # 로그인 정보를 받기 위해 cookie사용
    username = request.COOKIES.get('username')
    if username is not None:
        if User.objects.filter(username=username).exists():
            # 이미 존재하는 username일때만 로그인
            user = User.objects.filter(username=username).first()
            auth.login(request, user)
    return request

# 정의 : base
# 목적 : 웹 페이지 시작 페이지 로딩 
# 멤버함수 : logincheck
# 개발자 : 임수민, soomin910612@gmail.com
# 최종수정일 : 2022-02-22
def base(request):
    '''
    general page
    '''
    platforms = Platform.objects.filter(active=1) #get all platform info from db
    values = {
        'platforms': platforms
    }
    request = logincheck(request)
    return render(request, 'dataprocess/base.html',values)
    
# 정의 : daily
# 목적 : 데이터리포트 화면 로딩(get) 및 엑셀 export/import 기능(post)
# 멤버함수 : 
# 개발자 : 김민희, minheekim3@naver.com
# 최종수정일 : 2022-02-22
@csrf_exempt
def daily(request):
    if request.method == 'GET':
        # get을 통해 화면 로딩
        '''
        general page
        '''
        platforms = Platform.objects.filter(active=1) #get all platform info from db
        values = {
            'first_depth' : '데이터 리포트',
            'second_depth': '일별 리포트',
            'platforms': platforms
        }
        request = logincheck(request)
        return render(request, 'dataprocess/daily.html',values)
    else:
        # post를 통해 엑셀 export/import 기능 수행
        type = request.POST['type']
        if type == 'import':
            '''
            import from excel
            기능: 데이터리포트-크롤링데이터 엑셀 업로드
            '''
            platforms = Platform.objects.filter(active=1)  # get all platform info from db
            if not 'importData' in request.FILES:
                # 첨부파일이 없는 경우 에러메시지
                values = {
                    'first_depth' : '데이터 리포트',
                    'second_depth': '일별 리포트',
                    'platforms': platforms,
                    'alert': '파일을 첨부해주세요.'
                    }
                request = logincheck(request)
                return render(request, 'dataprocess/daily.html', values)
            import_file = request.FILES['importData']            
            excel_import_date = request.POST.get('excel_import_date', None)  # 0000-0-0 형태

            wb = openpyxl.load_workbook(import_file)
            sheets = wb.sheetnames
            worksheet = wb[sheets[0]]
            import_datareport(worksheet, excel_import_date)

            values = {
                'first_depth' : '데이터 리포트',
                'second_depth': '일별 리포트',
                'platforms': platforms,
                'alert': '저장되었습니다.'
                }
            request = logincheck(request)
            return render(request, 'dataprocess/daily.html',values)
        elif type == 'export':
            '''
            export to excel
            기능: 데이터리포트-크롤링데이터 엑셀 다운로드
            '''
            excel_export_type = request.POST.get('excel_export_days', None) # 누적 or 기간별
            excel_export_start_date = request.POST.get('excel_export_start_date', None) # 0000-0-0 형태
            excel_export_end_date = request.POST.get('excel_export_end_date', None) # 0000-0-0 형태
            book = export_datareport(excel_export_type, excel_export_start_date, excel_export_end_date)
            if excel_export_type == '누적':
                filename = "datareport %s.xlsx" % (excel_export_start_date)
            elif excel_export_type == '기간별':
                filename = "datareport %s~%s.xlsx" % (excel_export_start_date,excel_export_end_date)
            response = HttpResponse(content=save_virtual_workbook(book), content_type='application/vnd.ms-excel')
            response['Content-Disposition'] = 'attachment; filename='+filename
            return response
        elif type == 'import2':
            '''
            import2 from excel
            기능: 데이터리포트-수집정보 엑셀 업로드
            저장되는 DB table: collect_target_item, artist, platform
            '''
            platforms = Platform.objects.filter(active=1)  # get all platform info from db
            if not 'importData' in request.FILES:
                # 첨부파일이 없는 경우 에러메시지
                values = {
                    'first_depth' : '데이터 리포트',
                    'second_depth': '일별 리포트',
                    'platforms': platforms,
                    'alert': '파일을 첨부해주세요.'
                    }
                request = logincheck(request)
                return render(request, 'dataprocess/daily.html', values)
            import_file = request.FILES['importData']
            wb = openpyxl.load_workbook(import_file)
            sheets = wb.sheetnames
            worksheet = wb[sheets[0]]
            import_collects(worksheet)
            values = {
                'first_depth' : '데이터 리포트',
                'second_depth': '일별 리포트',
                'platforms': platforms,
                'alert': '저장되었습니다.'
                }
            request = logincheck(request)
            return render(request, 'dataprocess/daily.html',values)
        elif type == 'import3':
            '''
            import3 from excel
            기능: 데이터리포트-로그인정보 엑셀 업로드
            저장되는 DB table: auth_info
            '''
            platforms = Platform.objects.filter(active=1)  # get all platform info from db
            if not 'importData' in request.FILES:
                # 첨부파일이 없는 경우 에러메시지
                values = {
                    'first_depth' : '데이터 리포트',
                    'second_depth': '일별 리포트',
                    'platforms': platforms,
                    'alert': '파일을 첨부해주세요.'
                    }
                request = logincheck(request)
                return render(request, 'dataprocess/daily.html', values)
            import_file = request.FILES['importData']
            wb = openpyxl.load_workbook(import_file)
            sheets = wb.sheetnames
            worksheet = wb[sheets[0]]
            import_authinfo(worksheet)
            values = {
                'first_depth' : '데이터 리포트',
                'second_depth': '일별 리포트',
                'platforms': platforms,
                'alert': '저장되었습니다.'
                }
            request = logincheck(request)
            return render(request, 'dataprocess/daily.html',values)
        elif type == 'export_form':
            '''
            export_form to excel (데이터리포트-인증정보(collect_form), 로그인정보(login_form) 엑셀양식 다운로드)
            기능: media directory에 있는 파일을 전달
            '''
            file_name = request.POST['filename']
            if file_name == '':
                return None
            file_path = os.path.abspath("media/form")
            file_name = os.path.basename(f"media/form/{file_name}.xlsx")
            fs = FileSystemStorage(file_path)
            response = FileResponse(fs.open(file_name, 'rb'),
                                    content_type='application/vnd.ms-excel')
            response['Content-Disposition'] = f'attachment; filename="{file_name}"'
            return response

# 정의 : platform 페이지
# 목적 : 웹 페이지 플랫폼 관리 페이지 로딩 
# 멤버함수 : logincheck
# 개발자 : 임수민, soomin910612@gmail.com
# 최종수정일 : 2022-02-22
def platform(request):
     '''
     general page
     '''
     values = {
        'first_depth' : '플랫폼 관리',
        'second_depth': '플랫폼 관리'
    }
     request = logincheck(request)
     return render(request, 'dataprocess/platform.html',values)

# 정의 : artist 페이지
# 목적 : 웹 페이지 아티스트 중 데이터 URL 관리 페이지  
# 멤버함수 : logincheck
# 개발자 : 임수민, soomin910612@gmail.com
# 최종수정일 : 2022-02-22
def artist(request):
    artists = Artist.objects.all()
    # artist 이름 가나다순 정렬
    artists = list(artists.values())
    artists = sorted(sorted(artists, key=lambda c:c['name']), key=lambda c:0 if re.search('[ㄱ-힣]', c['name'][0]) else 1)
    values = {
      'first_depth' : '아티스트 관리',
      'second_depth': '데이터 URL 관리',
      'artists': artists
    }
    request = logincheck(request)
    return render(request, 'dataprocess/artist.html',values)

# 정의 : artist 페이지
# 목적 : 웹 페이지 아티스트 중 데이터 URL 관리 페이지 중 아티스트 추가 페이지
# 멤버함수 : logincheck
# 개발자 : 임수민, soomin910612@gmail.com
# 최종수정일 : 2022-02-22
@csrf_exempt
def artist_add(request):
    platforms = Platform.objects.all()
    values = {
      'first_depth' : '아티스트 관리',
      'second_depth': '데이터 URL 관리',
      'platforms' : platforms
    }
    request = logincheck(request)
    return render(request, 'dataprocess/artist_add.html',values)


#정의 : monitering 페이지
# 목적 : 웹 페이지 모니터링 중 모니터링 페이지 
# 멤버함수 : logincheck
# 개발자 : 임수민, soomin910612@gmail.com
# 최종수정일 : 2022-02-22
def monitering(request):
    platforms = Platform.objects.all()
    values = {
      'first_depth' : '모니터링 관리',
      'second_depth': '모니터링',
      'platforms' : platforms
    }
    request = logincheck(request)
    return render(request, 'dataprocess/monitering.html', values)

# 정의 : login 페이지
# 목적 : 로그인 페이지
# 멤버함수 : logincheck
# 개발자 : 김민희, minheekim3@naver.com
# 최종수정일 : 2022-02-22
def login(request):
    values = {
      'first_depth' : '로그인'
    }
    request = logincheck(request)
    return render(request, 'dataprocess/login.html',values)

# 정의 : 모니터링 URL 오류 관련 페이지네이션 기능
# 목적 : 모니터링 페이지에서 URL 오류 관련한 URL 정보들을 페이지네이션 된 데이터 조각으로 리턴
# 멤버함수 : get
# 개발자 : 임수민, soomin910612@gmail.com
# 최종수정일 : 2020-02-18
class ResultQueryView(ViewPaginatorMixin,APIView):
    def get(self, request):
        from_date_str = request.GET.get("fromdate", None)
        to_date_str = request.GET.get("todate", None)
        try:
            from_date_obj = datetime.datetime.strptime(from_date_str, '%Y-%m-%d')
            to_date_obj = datetime.datetime.strptime(to_date_str, '%Y-%m-%d')
        except Exception:
            return JsonResponse(status=400, data={"error": "Input Date Format Error"})
        page = request.GET.get('page',1)
        limit = 10
        day_diff = (to_date_obj - from_date_obj).days
        platforms = ["crowdtangle", "melon", "spotify", "tiktok", "twitter", "twitter2", "vlive", "weverse", "youtube"]
        error_details = [] # 전체 에러 디테일
        total_exec = 0
        for day in range(0, day_diff + 1):
            for platform in platforms:
                title_date = from_date_obj + datetime.timedelta(days=day)
                title_str = title_date.strftime("%Y-%m-%d")
                if production_env:
                    log_dir = f"../data/log/crawler/{platform}/{title_str}"
                else:
                    log_dir = f"./data/log/crawler/{platform}/{title_str}"
                if os.path.isdir(log_dir) is True:
                    file_list = os.listdir(log_dir)
                    for file_name in file_list:
                        task_id = file_name.split('.')[0]
                        task_result = get_task_result(task_id)
                        if task_result is not None and task_result == "started":
                            total_exec += 1
                        else:
                            platform_name, platform_artists, errors, error_infos = parse_logfile_for_error(f'{log_dir}/{file_name}')
                            if platform_name is not None:
                                for error_info in error_infos:
                                    if error_info['type'] == "400":
                                        artist_id = get_object_or_404(Artist,name = error_info['artist']).id
                                        platform_id = 0
                                        if platform == "crowdtangle": #instagram or facebook
                                            splited_url = error_info['url'].split('&') #split url by & 
                                            if "platform=facebook" in splited_url: #facebook case
                                                platform_id = Platform.objects.get(name = 'facebook').id
                                            else: #instagram case
                                                platform_id = Platform.objects.get(name = 'instagram').id
                                        else:
                                            platform_id = Platform.objects.get(name = error_info['platform']).id
                                        error_info['id'] = CollectTarget.objects.filter(artist = artist_id, platform = platform_id).first().id #collect target id
                                        error_details.append(error_info)
                                    else:
                                        continue

        return JsonResponse(data = {"data": self.paginate(error_details, page, limit)})

# 정의 : platform api
# 목적 : 플랫폼과 관련된 CRU api들
# 멤버함수 : get, post, put
# 개발자 : 김민희, minheekim3@naver.com
# 최종수정일 : 2022-02-22
class PlatformAPI(APIView):
    def get(self, request):
        '''
        Platform read api
        기능: platform table에 존재하는 모든 platform을 조회한다.
        '''
        try:
            platform_objects = Platform.objects.all()
            if platform_objects.exists():
                platform_objects_values = platform_objects.values()
                platform_datas = []
                for platform_value in platform_objects_values:
                    platform_datas.append(platform_value)
                return JsonResponse(data={'success': True, 'data': platform_datas})
            else:
                return JsonResponse(data={'success': True, 'data': []})
        except:
            return JsonResponse(status=400, data={'success': False})

    def post(self, request):
        '''
        Platform create api
        기능: platform, collect_target, schedule, collect_target_item table에 새로운 platform을 생성한다.
        '''
        try:
            platform_object = JSONParser().parse(request)
            platform_serializer = PlatformSerializer(data=platform_object)
            if platform_serializer.is_valid():
                # 1. platform 생성
                platform_serializer.save()

                # 2. 현재 존재하는 모든 artist에 대해 collect_target 생성
                # 위에서 생성한 platform_id 참조
                artist_objects = Artist.objects.all()
                artist_objects_values = artist_objects.values()
                for artist_objects_value in artist_objects_values:
                    collecttarget = CollectTarget(
                        platform_id = platform_serializer.data['id'],
                        artist_id = artist_objects_value['id']
                        )
                    collecttarget.save()
                    # 3. 해당 collect_target에 대한 schedule 생성
                    schedule_object = Schedule.objects.filter(collect_target_id = collecttarget.id).first()
                    schedule_data = {
                            'collect_target': collecttarget.id,
                            'schedule_type': 'daily',
                            'active': True,
                        }
                    schedule_serializer = ScheduleSerializer(schedule_object, data=schedule_data)
                    if schedule_serializer.is_valid():
                        schedule_serializer.save()
                    #4. 생성한 collect target에 대해 collect_target_item(수집항목)들 생성
                    collecttarget_object = CollectTarget.objects.filter(platform = platform_serializer.data['id'],
                            artist = artist_objects_value['id'])
                    collecttarget_object = collecttarget_object.values()[0]
                    for collect_item in platform_object['collect_items']:
                        collect_item = CollectTargetItem(
                            collect_target_id=collecttarget_object['id'],
                            target_name=collect_item['target_name'],
                            xpath=collect_item['xpath']
                        )
                        collect_item.save()
                return JsonResponse(data={'success': True, 'data': platform_serializer.data}, status=status.HTTP_201_CREATED)
            return JsonResponse(data={'success': False,'data': platform_serializer.errors}, status=400)
        except:
            return JsonResponse(data={'success': False}, status=400)

    def put(self, request):
        '''
        Platform update api
        기능: platform, schedule table을 수정한다.
        '''
        try:
            platform_list = JSONParser().parse(request)
            for platform_object in platform_list:
                platform_data = Platform.objects.filter(pk=platform_object['id']).first()
                if platform_data is None:
                    # 없는 경우 platform 신규 생성
                    platform_serializer = PlatformSerializer(data=platform_object)
                    if platform_serializer.is_valid():
                        platform_serializer.save()
                else:
                    # 있는 경우 기존 platform 업데이트
                    data = PlatformSerializer(platform_data).data
                    past_name = data['name']
                    past_url = data['url']
                    cur_name = platform_object['name']
                    cur_url = platform_object['url']
                    platform_serializer = PlatformSerializer(platform_data, data=platform_object)
                    if platform_serializer.is_valid():
                        if past_name != cur_name:
                            userlogger.info(f"[CHANGE]: {past_name} -> {cur_name}")
                        if past_url != cur_url:
                            userlogger.info(f"[CHANGE]: {past_url} -> {cur_url}")
                        platform_serializer.save()
                collecttarget_objects = CollectTarget.objects.filter(platform_id = platform_serializer.data['id'])
                # 해당 platform과 연관된 schedule들의 active 수정
                # 비활성이 우선이기 때문에 artist가 비활성이라면 넘어가고 활성일 경우에만 새로운 active값으로 수정
                if collecttarget_objects.exists():
                    collecttarget_values = collecttarget_objects.values()
                    for collecttarget_value in collecttarget_values:
                        artist_object = Artist.objects.get(pk = collecttarget_value['artist_id'])
                        if artist_object.active == True:
                            schedule_objects = Schedule.objects.filter(collect_target_id = collecttarget_value['id'])
                            schedule_objects.update(active = platform_object['active'])
            return JsonResponse(data={'success': True}, status=status.HTTP_201_CREATED)
        except:
            return JsonResponse(data={'success': False}, status=400)

# 정의 : artist api
# 목적 : 아티스트와 관련된 CRU api들
# 멤버함수 : get, post, put
# 개발자 : 김민희, minheekim3@naver.com
# 최종수정일 : 2022-02-22
class ArtistAPI(APIView):
    def get(self, request):
        '''
        Artist read api
        기능: artist table에 존재하는 모든 artist들을 조회한다.
        '''
        try:
            artist_objects = Artist.objects.all()
            if artist_objects.exists():
                # artist 이름 가나다순 정렬
                artist_objects_values = list(artist_objects.values())
                artist_objects_values = sorted(sorted(artist_objects_values, key=lambda c:c['name']), key=lambda c:0 if re.search('[ㄱ-힣]', c['name'][0]) else 1)
                artist_datas = []
                for artist_value in artist_objects_values:
                    artist_datas.append(artist_value)
                return JsonResponse(status=200,data={'success': True, 'data': artist_datas})
            else:
                return JsonResponse(data={'success': True, 'data': []})
        except:
            return JsonResponse(status=400, data={'success': False})

    def post(self, request):
        '''
        Artist create api
        기능: artist, collect_target, schedule, collect_target_item table에 새로운 artist를 생성한다.
        '''
        try:
            artist_object = JSONParser().parse(request)
            artist_serializer = ArtistSerializer(data=artist_object)
            if artist_serializer.is_valid():
                # 1. artist 신규 생성
                artist_serializer.save()
                # 2. 현재 존재하는 모든 platform에 대해 collect_target 생성
                # 위에서 생성한 artist_id 참조
                for obj in artist_object['urls']:
                    platform_id = Platform.objects.get(name = obj['platform_name']).id
                    artist_id = artist_serializer.data['id']
                    target_url = obj['url1']
                    target_url_2 = obj['url2']
                    if target_url == "" and target_url_2 == "":
                        continue
                    collecttarget = CollectTarget(
                        platform_id=platform_id,
                        artist_id=artist_id,
                        target_url=target_url,
                        target_url_2=target_url_2
                    )
                    collecttarget.save()
                    # 3. 해당 collect_target에 대한 schedule 생성
                    # 기존 platform의 daily schedule과 똑같은 execute_time 사용
                    schedule_object = Schedule.objects.filter(collect_target_id = collecttarget.id).first()
                    collecttarget_objects = CollectTarget.objects.filter(artist_id = artist_id)
                    collecttarget_objects = collecttarget_objects.values()
                    execute_time = None
                    for collecttarget_object in collecttarget_objects:
                        schedule_objects = Schedule.objects.filter(schedule_type = 'daily', collect_target_id = collecttarget_object['id']).values()
                        if schedule_objects.exists():
                            execute_time = schedule_objects[0]['execute_time']
                            break
                    schedule_data = {
                            'collect_target': collecttarget.id,
                            'schedule_type': 'daily',
                            'active': True,
                            'execute_time': execute_time
                        }
                    schedule_serializer = ScheduleSerializer(schedule_object, data=schedule_data)
                    if schedule_serializer.is_valid():
                        schedule_serializer.save()
                    # 4. 현재 플랫폼의 존재하는 조사항목들을 default로 collect_target_item 생성
                    # 해당 플랫폼을 참조하는 첫번째 collect_target의 collect_target_item들 사용
                    collecttarget_objects = CollectTarget.objects.filter(artist_id = artist_id, platform_id = platform_id)
                    collecttarget_object = collecttarget_objects.values().first()
                    first_collect_target = CollectTarget.objects.filter(platform_id=platform_id).first()
                    collect_target_items_objects = CollectTargetItem.objects.filter(collect_target = first_collect_target.id)
                    for collect_target_items_object in collect_target_items_objects:
                        collecttargetitem_serializer = CollectTargetItemSerializer(data={
                            'collect_target': collecttarget_object['id'],
                            'target_name': collect_target_items_object.target_name,
                            'xpath': ""
                        })
                    if collecttargetitem_serializer.is_valid():
                        collecttargetitem_serializer.save()
                return JsonResponse(data={'success': True}, status=status.HTTP_201_CREATED)
            return JsonResponse(data={'success': False,'data': artist_serializer.errors}, status=400)
        except:
            return JsonResponse(data={'success': False}, status=400)

    def put(self, request):
        '''
        Artist update api
        기능: artist, schedule table을 수정한다.
        '''
        try:
            artist_list = JSONParser().parse(request)
            for artist_object in artist_list:
                artist_data = Artist.objects.get(id=artist_object["id"])
                data = ArtistSerializer(artist_data).data
                past_name = data['name']
                past_num = data['member_num']
                past_agency = data['agency']
                cur_name = artist_object['name']
                cur_num = artist_object['member_num']
                cur_agency = artist_object['agency']
                artist_serializer = ArtistSerializer(artist_data, data=artist_object)
                if artist_serializer.is_valid():
                    if past_name != cur_name:
                        userlogger.info(f"[CHANGE]: {past_name} -> {cur_name}")
                    if past_num != cur_num:
                        userlogger.info(f"[CHANGE]: {past_num} -> {cur_num}")
                    if past_agency != cur_agency:
                        userlogger.info(f"[CHANGE]: {past_agency} -> {cur_agency}")
                    # artist 수정
                    artist_serializer.save()
                else:
                    return JsonResponse(data={'success': False, 'data': artist_serializer.errors}, status=400)
                collecttarget_objects = CollectTarget.objects.filter(artist_id = artist_serializer.data['id'])
                # artist와 연관된 schedule들 수정
                # 비활성이 우선이기 때문에 platform이 비활성이라면 넘어가고 활성일 경우에만 새로운 active값으로 수정
                if collecttarget_objects.exists():
                    collecttarget_values = collecttarget_objects.values()
                    for collecttarget_value in collecttarget_values:
                        platform_object = Platform.objects.get(pk = collecttarget_value['platform_id'])
                        if platform_object.active == True:
                            schedule_objects = Schedule.objects.filter(collect_target_id = collecttarget_value['id'])
                            schedule_objects.update(active = artist_object['active'])
            return JsonResponse(data={'success': True}, status=status.HTTP_201_CREATED)
        except:
            return JsonResponse(data={'success': False}, status=400)


# 정의 : 아티스트별 플랫폼 정보 URL 불러오기/수정 
# 목적 : 웹 페이지 아티스트 중 데이터 URL 관리 페이지에서 아티스트 클릭 시 나오는 플랫폼별 URL을 불러오고 수정한다.   
# 멤버함수 : get, put
# 개발자 : 김민희, minheekim3@naver.com(get)/ 임수민, soomin910612@gmail.com(put)
# 최종수정일 : 2022-02-22
class PlatformOfArtistAPI(APIView):
    def get(self, request):
        '''
        Platform of Artist read api
        기능: 해당 아티스트가 수집하고자하는 플랫폼들을 조회한다.
        '''
        try:
            artist = request.GET.get('artist', None)
            # 해당 artist 찾기
            artist_object = Artist.objects.filter(name = artist)
            artist_object = artist_object.values()[0]
            # 해당 artist_id를 참조하는 collect_target들 조회
            collecttarget_objects = CollectTarget.objects.filter(artist_id=artist_object['id'])
            if collecttarget_objects.exists():
                collecttarget_objects_values = collecttarget_objects.values()
                platform_datas = []
                for collecttarget_value in collecttarget_objects_values:
                    platform_object = Platform.objects.get(pk=collecttarget_value['platform_id'])
                    platform_datas.append({
                        'artist_id':artist_object['id'],
                        'platform_id' : collecttarget_value['platform_id'],
                        'id': collecttarget_value['id'],
                        'name': platform_object.name,
                        'target_url':collecttarget_value['target_url'],
                        'target_url_2': collecttarget_value['target_url_2']
                    })
                return JsonResponse(data={'success': True, 'data': platform_datas})
            else:
                return JsonResponse(data={'success': True, 'data': []})
        except:
            return JsonResponse(status=400, data={'success': False})

    def put(self, request):
        '''
        Platform of Artist update api
        기능: collect_target table을 수정한다.
        '''
        try:
            collecttarget_list = JSONParser().parse(request)
            data = ''
            for collecttarget_object in collecttarget_list:
                if collecttarget_object['type'] == 'artist-platform-update':
                    CollectTarget.objects.filter(pk=collecttarget_object['id']).update(target_url=collecttarget_object['target_url'])
                    CollectTarget.objects.filter(pk=collecttarget_object['id']).update(target_url_2=collecttarget_object['target_url_2'])
                else:
                    target_obj = CollectTarget.objects.filter(pk=collecttarget_object['id'])
                    target_obj_value = target_obj.values()[0]
                    data = collecttarget_object['new_target_url']

                    if target_obj_value['target_url'] == collecttarget_object['old_target_url']:
                        target_obj.update(target_url = collecttarget_object['new_target_url'])
                    elif target_obj_value['target_url_2'] == collecttarget_object['old_target_url']:
                        target_obj.update(target_url_2 = collecttarget_object['new_target_url'])
                    else:
                        data = ''
            
            return JsonResponse(data={'success': True,'data':data}, status=status.HTTP_201_CREATED)
              
        except:
            return JsonResponse(data={'success': False}, status=400)

# 정의 : collecttargetitem(조사항목) api
# 목적 : 조사항목과 관련된 RUD api들
# 멤버함수 : get, put, delete
# 개발자 : 김민희, minheekim3@naver.com / 양승찬, uvzone@naver.com(put)
# 최종수정일 : 2022-02-22
    def get(self, request):
        '''
        CollectTargetItem read api
        기능: 아티스트-플랫폼이 수집하고자하는 조사항목들을 조회한다.
        '''
        try:
            artist = request.GET.get('artist', None)
            platform = request.GET.get('platform', None)
            # 해당 artist, platform 찾기
            artist_object = Artist.objects.filter(name=artist).first()
            platform_object = Platform.objects.filter(name=platform).first()
            # 해당 artist와 platform을 가지는 collect_target 가져오기
            collecttarget_object = CollectTarget.objects.filter(artist_id=artist_object.id, platform_id=platform_object.id)
            if collecttarget_object.exists():
                collecttarget_object = collecttarget_object.first()
                collecttargetitems_datas = []
                # 조사항목들 조회
                collecttargetitmes_objects = CollectTargetItem.objects.filter(collect_target_id=collecttarget_object.id)
                collecttargetitmes_values = collecttargetitmes_objects.values()
                for collecttargetitmes_value in collecttargetitmes_values:
                    collecttargetitems_datas.append(collecttargetitmes_value)
                # schedule 조회
                schedule_object = Schedule.objects.filter(collect_target_id = collecttarget_object.id)
                if schedule_object.exists():
                    schedule_type = schedule_object.values()[0]['schedule_type']
                else:
                    schedule_type = 'daily'
                return JsonResponse(data={'success': True, 'data': {'items':collecttargetitems_datas, 'schedule_type': schedule_type}})
            else:
                return JsonResponse(data={'success': True, 'data': {'items':[],'schedule_type':'daily'}})
        except:
            return JsonResponse(status=400, data={'success': False})

    # TODO : 스케줄 수정 시 해당하는 아티스트가 없으면 스케줄링에서 삭제하기
    def put(self, request):
        '''
        CollectTargetItem update api
        기능: collect_target_item, schedule table을 수정한다.
        '''
        try:
            collecttargetitem = JSONParser().parse(request)
            artist = collecttargetitem['artist']
            platform = collecttargetitem['platform']
            schedule_type = collecttargetitem['schedule_type']
            collecttargetitem_list = collecttargetitem['items']
            artist_object = Artist.objects.filter(name = collecttargetitem['artist']).first()
            platform_object = Platform.objects.filter(name = collecttargetitem['platform']).first()
            collecttarget_object = CollectTarget.objects.filter(artist_id=artist_object.id, platform_id=platform_object.id).first()
            for collecttargetitem_object in collecttargetitem_list:
                collecttargetitem_data = CollectTargetItem.objects.filter(id=collecttargetitem_object['id'],
                                                                          target_name=collecttargetitem_object['target_name'], xpath=collecttargetitem_object['xpath']).first()
                # 조사항목이 없다면 신규 생성
                if collecttargetitem_data is None:
                    collecttargetitem_serializer = CollectTargetItemSerializer(data={
                        'collect_target': collecttarget_object.id,
                        'target_name': collecttargetitem_object['target_name'],
                        'xpath': collecttargetitem_object['xpath']
                    })
                    if collecttargetitem_serializer.is_valid():
                        collecttargetitem_serializer.save()
                    else:
                        return JsonResponse(data={'success': False}, status=400)
                # 조사항목이 있으면 업데이트
                else:
                    collecttargetitem_serializer = CollectTargetItemSerializer(collecttargetitem_data, data=collecttargetitem_object)
                    if collecttargetitem_serializer.is_valid():
                        collecttargetitem_serializer.save()
                        userlogger.debug(f"{artist} - {platform} - {schedule_type}: ")
                    else:
                        return JsonResponse(data={'success': False,'data': collecttargetitem_serializer.errors}, status=400)
            # 해당 schedule 업데이트
            # platform내의 다른 schedule과 execute_time, period 동일시
            execute_time = None # 시작 시간
            period = None # 주기
            collecttarget_objects = CollectTarget.objects.filter(platform_id = platform_object.id)
            collecttarget_objects = collecttarget_objects.values()
            for collecttarget_value in collecttarget_objects:
                schedule_objects = Schedule.objects.filter(schedule_type = collecttargetitem['schedule_type'], collect_target_id = collecttarget_value['id'])
                if schedule_objects.exists():
                    execute_time = schedule_objects.values()[0]['execute_time']
                    period = schedule_objects.values()[0]['period']
                    break
            Schedule.objects.filter(collect_target_id = collecttarget_object.id).update(
                    schedule_type = schedule_type, execute_time = execute_time, period = period)

            # Schedule 초기화 여부를 확인하는 부분
            prev_schedule_type = 'daily'
            if schedule_type == 'daily':
                prev_schedule_type = 'hour'
            no_schedule_collect_targets = True
            for collecttarget_value in collecttarget_objects:
                prev_schedule_objects = Schedule.objects.filter(schedule_type = prev_schedule_type, collect_target_id = collecttarget_value['id'])
                if prev_schedule_objects.exists():
                    no_schedule_collect_targets = False
                    break
            if no_schedule_collect_targets:
                try:
                    schedule_name = f"{platform}_{prev_schedule_type}_crawling"
                    found_schedule = PeriodicTask.objects.get(name=schedule_name)
                    found_schedule.delete()
                except PeriodicTask.DoesNotExist:
                    found_schedule = None
            return JsonResponse(data={'success': True}, status=status.HTTP_201_CREATED)
        except:
            return JsonResponse(data={'success': False}, status=400)

    def delete(self, request):
        '''
        CollectTargetItem delete api
        기능: collect_target_item table에서 조사항목을 삭제한다.
        '''
        try:
            delete_id = JSONParser().parse(request)['id']
            obj = CollectTargetItem.objects.filter(id = delete_id)
            obj.delete()
            return JsonResponse(data={'success': True}, status=status.HTTP_200_OK)
        except:
            return JsonResponse(data={'success': False}, status=400)

# 정의 : 데이터 리포트 일별 데이터 불러오기 및 수정
# 목적 : 크롤링 된 데이터를 불러오고 수정한다. 크롤링 데이터 형식은 json 이다. 
# 멤버함수 : get, post
# 개발자 : 김민희,  minheekim3@naveer.com(get)/ 임수민, soomin910612@gmail.com(post)
# 최종수정일 : 2022-02-22
class DataReportAPI(APIView):
    def get(self, request):
        '''
        Data-Report read api
        기능: 원하는 날짜의 데이터 리포트 일별/시간별 데이터를 조회한다.
        '''
        platform = request.GET.get('platform', None)
        type = request.GET.get('type', None)
        start_date = request.GET.get('start_date', None)
        end_date = request.GET.get('end_date', None)

        # artist name
        artist_objects = Artist.objects.filter(active=1)
        artist_objects_values = list(artist_objects.values())
        # artist name 가나다순 정렬
        artist_objects_values = sorted(sorted(artist_objects_values, key=lambda c:c['name']), key=lambda c:0 if re.search('[ㄱ-힣]', c['name'][0]) else 1)
        artist_list = []
        for a in artist_objects_values:
            artist_list.append(a['name'])

        #platform target names
        platform_id = Platform.objects.get(name = platform).id
        collecttargets = CollectTarget.objects.filter(platform = platform_id)
        collecttargets = collecttargets.values()
        platform_list = set()
        for collecttarget in collecttargets:
            platform_objects = CollectTargetItem.objects.filter(collect_target_id = collecttarget['id'])
            platform_objects_values = platform_objects.values()
            for p in platform_objects_values:
                if p['target_name'] in platform_list:
                    continue
                platform_list.add(p['target_name'])
        platform_list = list(platform_list)
        #플랫폼 헤더 정보 순서와 db 칼럼 저장 순서 싱크 맞추기
        platform_header = []
        objects = CollectData.objects.filter(collect_items__platform=platform)
        objects_values = objects.values()
        if len(objects_values) > 0:
            key_list = list(objects_values[0]['collect_items'].keys())
            for key in key_list:
                if key in platform_list:
                    platform_header.append(key)
                else:
                    continue
        else:
            platform_header = platform_list

        try:
            if type == '누적':
                '''
                누적 데이터 조회
                '''
                start_date_dateobject = datetime.datetime.strptime(start_date, "%Y-%m-%d")
                start_date_string = start_date_dateobject.strftime("%Y-%m-%d")
                # 해당날짜에 데이터가 존재하는지 저장하는 변수 check
                check = False
                crawling_artist_list = []
                objects = CollectData.objects.filter(collect_items__platform=platform)
                objects_value = objects.values()
                for val in objects_value:
                    crawling_artist_list.append(val["collect_items"]["artist"])
                filter_datas = []
                for artist in artist_list:
                    filter_objects = CollectData.objects.filter(collect_items__artist=artist, collect_items__platform=platform,
                        collect_items__reserved_date = start_date_string)
                    if filter_objects.exists():
                        check = True
                        # 같은 날짜에 여러개 있을 때 가장 앞의 데이터를 가져온다
                        filter_value = filter_objects.values()[0]
                        filter_datas.append(filter_value['collect_items'])
                # 해당날짜에 데이터가 하나라도 있을 때
                if check:
                    return JsonResponse(data={'success': True, 'data': filter_datas, 'artists': artist_list, 'platform': platform_header,'crawling_artist_list': crawling_artist_list})
                # 해당날짜에 데이터가 하나도 없을 때
                else:
                    crawling_artist_list = []
                    objects = CollectData.objects.filter(collect_items__platform=platform)
                    objects_value = objects.values()
                    for val in objects_value:
                        crawling_artist_list.append(val['collect_items']['artist'])
                    return JsonResponse(status=200, data={'success': True, 'data': 'no data', 'artists': artist_list, 'platform': platform_header,
                                                        'crawling_artist_list': crawling_artist_list})

            elif type == '기간별':
                '''
                기간별 데이터 조회
                '''
                # 하루 전날의 date를 가져옴
                start_date_dateobject = datetime.datetime.strptime(start_date, "%Y-%m-%d").date() - datetime.timedelta(1)
                end_date_dateobject = datetime.datetime.strptime(end_date, "%Y-%m-%d").date()
                start_date_string = start_date_dateobject.strftime("%Y-%m-%d")
                end_date_string = end_date_dateobject.strftime("%Y-%m-%d")
                # 해당날짜에 데이터가 존재하는지 저장하는 변수 check
                check = False
                crawling_artist_list = []
                objects = CollectData.objects.filter(collect_items__platform=platform)
                objects_value = objects.values()
                for val in objects_value:
                    crawling_artist_list.append(val['collect_items']['artist'])
                filter_datas_total = []
                for artist in artist_list:
                    filter_objects_start = CollectData.objects.filter(collect_items__artist=artist, collect_items__platform=platform,
                            collect_items__reserved_date = start_date_string)
                    filter_objects_end = CollectData.objects.filter(collect_items__artist=artist, collect_items__platform=platform,
                            collect_items__reserved_date = end_date_string)
                    # 양끝날짜 데이터 모두 존재할 때
                    if filter_objects_start.exists() and filter_objects_end.exists():
                        check = True
                        filter_objects_start_value = filter_objects_start.values()[0]
                        filter_objects_start_value = filter_objects_start_value['collect_items']
                        # id랑 artist, date 빼고 보내주기
                        data_json = {}
                        filter_objects_end_value = filter_objects_end.values()[0]
                        filter_objects_end_value = filter_objects_end_value['collect_items']
                        for field_name in filter_objects_start_value.keys():
                            if field_name != 'id' and field_name != 'artist' and field_name != 'user_created' and field_name != 'recorded_date' and field_name != 'platform' and field_name != 'url' and field_name != 'url1' and field_name != 'url2' and field_name != 'reserved_date' and field_name != 'updated_dt':
                                if filter_objects_end_value[field_name] is not None and filter_objects_start_value[field_name] is not None:
                                    data_json[field_name] = int(filter_objects_end_value[field_name]) - int(filter_objects_start_value[field_name])
                                    data_json[field_name+'_end'] = int(filter_objects_end_value[field_name])
                                elif filter_objects_end_value[field_name] is not None:  # 앞의 날짜를 0으로 처리한 형태
                                    data_json[field_name] = filter_objects_end_value[field_name]
                                    data_json[field_name+'_end'] = int(filter_objects_end_value[field_name])
                                else: # 앞의 날짜가 없다면 0으로 보내기
                                    data_json[field_name] = 0
                                    data_json[field_name+'_end'] = filter_objects_end_value[field_name]
                                data_json[field_name+'_end'] = filter_objects_end_value[field_name]
                            else:  # 숫자 아닌 다른 정보들(user_created 등)
                                data_json[field_name] = filter_objects_start_value[field_name]
                        filter_datas_total.append(data_json)
                    elif not filter_objects_start.exists() and filter_objects_end.exists():
                        # 시작날짜의 데이터가 존재하지 않고 끝날짜의 데이터만 존재할 때
                        # 시작날짜: 0으로 해서 계산 -> 끝날짜 데이터 자체를 보냄
                        check = True
                        filter_objects_end_value = filter_objects_end.values()[0]
                        filter_objects_end_value = filter_objects_end_value['collect_items']
                        filter_datas_total.append(filter_objects_end_value)
                if check: # 양끝 모두 존재 or 끝날짜만 존재
                    return JsonResponse(data={'success': True, 'data': filter_datas_total, 'artists': artist_list, 'platform': platform_header,'crawling_artist_list':crawling_artist_list})
                else: # 끝날짜의 데이터가 아예 존재하지 않을 때
                    return JsonResponse(status=400, data={'success': False, 'data': end_date})
            else:
                '''
                누적도 기간별도 아닌 경우(에러처리)
                '''
                start_date_dateobject = datetime.datetime.strptime(start_date, "%Y-%m-%d")
                start_date_string = start_date_dateobject.strftime("%Y-%m-%d")
                crawling_artist_list = []
                objects = CollectData.objects.filter(collect_items__platform=platform)
                objects_value = objects.values()
                for val in objects_value:
                    crawling_artist_list.append(val['collect_items']['artist'])
                objects = CollectData.objects.filter(collect_items__platform=platform,
                            collect_items__reserved_date = start_date_string)
                if objects.exists():
                    platform_queryset_values = objects.values()
                    platform_datas = []
                    for queryset_value in platform_queryset_values:
                        platform_datas.append(queryset_value['collect_items'])
                    return JsonResponse(data={'success': True, 'data': platform_datas, 'artists': artist_list, 'platform': platform_header,'crawling_artist_list':crawling_artist_list})
                else:
                    return JsonResponse(status=400, data={'success': False, 'data': start_date})
        except:
            return JsonResponse(status=400, data={'success': False, 'data': start_date})
    
    def post(self, request):
        '''
        Data-Report update api
        기능: 누적 데이터를 수정한다.
        '''
        update_data_object = JSONParser().parse(request)
        start_date = update_data_object[len(update_data_object)-1]['start_date']
        platform = update_data_object[len(update_data_object)-1]['platform_name']
        
        # artist name
        artist_objects = Artist.objects.filter(active=1)
        artist_objects_values = list(artist_objects.values())
        artist_objects_values = sorted(sorted(artist_objects_values, key=lambda c:c['name']), key=lambda c:0 if re.search('[ㄱ-힣]', c['name'][0]) else 1)
        artist_list = []
        for a in artist_objects_values:
            artist_list.append(a['name'])

        # crawled artist list
        a_objects = CollectData.objects.filter(collect_items__platform=platform)
        a_objects_values = a_objects.values()
        a_list = []
        for val in a_objects_values:
            a_list.append(val['collect_items']['artist'])

        #platform target names
        platform_id = Platform.objects.get(name = platform).id
        collecttargets = CollectTarget.objects.filter(platform = platform_id)
        collecttargets = collecttargets.values()
        platform_list = set()
        for collecttarget in collecttargets:
            platform_objects = CollectTargetItem.objects.filter(collect_target_id = collecttarget['id'])
            platform_objects_values = platform_objects.values()
            for p in platform_objects_values:
                if p['target_name'] in platform_list:
                    continue
                platform_list.add(p['target_name'])
        platform_list = list(platform_list)

        #플랫폼 헤더 정보 순서와 db 칼럼 저장 순서 싱크 맞추기
        platform_header = []
        objects = CollectData.objects.filter(collect_items__platform=platform)
        objects_values = objects.values()
        if len(objects_values) > 0:
            key_list = list(objects_values[0]['collect_items'].keys())
            for key in key_list:
                if key in platform_list:
                    platform_header.append(key)
                else:
                    continue
        else:
            platform_header = platform_list

        try:
            start_date_dateobject = datetime.datetime.strptime(start_date, "%Y-%m-%d")
            start_date_string = start_date_dateobject.strftime("%Y-%m-%d")
            for index, element in enumerate(update_data_object):
                if index == len(update_data_object)-1:
                    break
                artist_object = Artist.objects.filter(name=element['artist'])
                artist_object = artist_object.values()[0]
                platform_object = Platform.objects.filter(name=platform)
                platform_object = platform_object.values()[0]
                collecttarget_object = CollectTarget.objects.filter(platform = platform_object['id'],
                                        artist = artist_object['id'])
                collecttarget_object = collecttarget_object.values()[0]
                CollectData.objects.update_or_create(
                        collect_target_id = collecttarget_object['id'],
                        collect_items__reserved_date = start_date_string,
                        # collect_items = element,
                        # 바뀌는 값
                        defaults = {'collect_items': element}
                )
                    
            artist_set = set()
            filter_objects = CollectData.objects.filter(collect_items__platform=platform,
                            collect_items__reserved_date = start_date_string)
            if filter_objects.exists():
                filter_objects_values=filter_objects.values()
                filter_datas=[]

                crawling_artist_list = set()
                platform_filter_objects = CollectData.objects.filter(collect_items__platform=platform)
                objects_value = platform_filter_objects.values()
                for val in objects_value:
                    val = val['collect_items']
                    # 각 아티스트가 한번만 들어가도록
                    if val['artist'] in crawling_artist_list:
                        continue
                    crawling_artist_list.add(val['artist'])
                crawling_artist_list = list(crawling_artist_list)
                for filter_value in filter_objects_values:
                    filter_value = filter_value['collect_items']
                    # 각 아티스트당 하나의 데이터만 들어가도록
                    if filter_value['artist'] in artist_set:
                        continue
                    filter_datas.append(filter_value)
                return JsonResponse(data={'success': True, 'data': filter_datas,'artists':artist_list,'platform':platform_header,'crawling_artist_list':crawling_artist_list})
            else:
                # 존재하지 않을 때 -> 플랫폼 전체 데이터를 보자
                filter_objects = CollectData.objects.filter(collect_items__platform=platform)
                crawling_artist_list = set()
                objects_value = filter_objects.values()
                for val in objects_value:
                    val = val['collect_items']
                    # 각 아티스트가 한번만 들어가도록
                    if val['artist'] in crawling_artist_list:
                        continue
                    crawling_artist_list.add(val['artist'])
                crawling_artist_list = list(crawling_artist_list)
                # datename = "%s-%s-%s"%(start_date_dateobject.year, start_date_dateobject.month, start_date_dateobject.day)
                return JsonResponse(status=200, data={'success': True, 'data': 'no data', 'artists': artist_list, 'platform': platform_header, 'crawling_artist_list': crawling_artist_list})
        except:
            return JsonResponse(status=400, data={'success': False})

# 정의 : schedule(스케줄) api
# 목적 : 조사항목과 관련된 RU api들
# 멤버함수 : get, put
# 개발자 : 김민희, minheekim3@naver.com
# 최종수정일 : 2022-02-22
class ScheduleAPI(APIView):
    def get(self, request):
        '''
        Schedule read api
        기능: 시간별 또는 일별을 입력받아 플랫폼마다의 schedule 정보를 조회한다.
        현재는 '시간별'일 경우에만 api 사용중임
        '''
        type = request.GET.get('type', None) # 시간별 or 일별
        try:
            if type == '시간별':
                # 해당 플랫폼에 시간별인 아티스트들 가져오기
                platform_objects = Platform.objects.all()
                platform_objects = platform_objects.values()
                hourly_list = []
                for platform_object in platform_objects:
                    period = None
                    execute_time = None
                    hour_artist_list = []
                    collecttarget_objects = CollectTarget.objects.filter(platform_id = platform_object['id'])
                    collecttarget_objects = collecttarget_objects.values()
                    for collecttarget_object in collecttarget_objects:
                        # 특정 platform의 시간별 artist 조회
                        schedule_objects = Schedule.objects.filter(schedule_type = 'hour', collect_target_id = collecttarget_object['id'])
                        if schedule_objects.exists():
                            period = schedule_objects.values()[0]['period']
                            execute_time = schedule_objects.values()[0]['execute_time']
                            artist = Artist.objects.get(pk = collecttarget_object['artist_id'])
                            hour_artist_list.append(artist.name)
                    hour_artist_list = sorted(sorted(hour_artist_list), key=lambda c:0 if re.search('[ㄱ-힣]', c[0]) else 1)
                    hourly_list.append({
                        'platform': platform_object['name'],
                        'artists': hour_artist_list,
                        'period': period,
                        'execute_time': execute_time
                    })
                return JsonResponse(data={'success': True, 'data': hourly_list})
            elif type == "일별":
                # 현재 "일별"일 경우 이 api 사용하지 않고 있음
                hourly_list = []
                return JsonResponse(data={'success': True, 'data': hourly_list})
        except:
            return JsonResponse(status=400, data={'success': False})

    def put(self, request):
        '''
        Schedule update api
        기능: schedule table을 업데이트한다.
        '''
        try:
            new_schedule = JSONParser().parse(request)
            schedule_type = new_schedule['schedule_type']
            # platform: facebook, instagram인 경우 crowdtangle라는 이름으로 프론트에서 전달
            # 함께 schedule 관리되어야함
            if new_schedule['platform'] == 'crowdtangle':
                platform_list = ['facebook', 'instagram']
            else:
                platform_list = [new_schedule['platform']]
            for platform_name in platform_list:
                platform_objects = Platform.objects.filter(name = platform_name)
                if platform_objects.exists():
                    collecttarget_objects = CollectTarget.objects.filter(platform_id = platform_objects.values()[0]['id'])
                    collecttarget_objects = collecttarget_objects.values()
                    for collecttarget_object in collecttarget_objects:
                        # 해당 platform과 연결된 schedule들의 모든 값 수정
                        if schedule_type == 'hour':
                            schedule_objects = Schedule.objects.filter(collect_target_id = collecttarget_object['id'], schedule_type = 'hour')
                            if schedule_objects.exists():
                                schedule_objects.update(period=datetime.time(new_schedule['period'],0,0), execute_time = datetime.time(0,new_schedule['execute_time_minute'],0))
                        elif schedule_type == 'daily':
                            schedule_objects = Schedule.objects.filter(collect_target_id = collecttarget_object['id'], schedule_type = 'daily')
                            if schedule_objects.exists():
                                schedule_objects.update(execute_time = datetime.time(new_schedule['execute_time_hour'],new_schedule['execute_time_minute'],0))
            return JsonResponse(data={'success': True}, status=status.HTTP_201_CREATED)
        except:
            return JsonResponse(data={'success': False}, status=400)
