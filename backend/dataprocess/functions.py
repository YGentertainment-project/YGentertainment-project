import datetime
import re
from dateutil.parser import parse
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, fonts
from openpyxl.styles.alignment import Alignment
from openpyxl.utils import get_column_letter
from config.models import CollectTargetItem, Schedule, AuthInfo
from dataprocess.models import Platform, Artist, CollectTarget, CollectData

from config.serializers import CollectTargetItemSerializer, ScheduleSerializer, AuthInfoSerializer
from dataprocess.serializers import ArtistSerializer, PlatformSerializer, CollectTargetSerializer

# 정의 : collect_data table을 조회하여 특정 플랫폼, 아티스트의 크롤링 데이터 list를 반환한다.
# 이때 반환하는 list는 엑셀파일 내의 조사항목 헤더와 싱크가 맞아야 한다.
# 목적 : 크롤링데이터 엑셀 다운로드(export)를 위해 크롤링 데이터 반환
# 멤버함수 : 
# 개발자 : 김민희, minheekim3@naver.com
# 최종수정일 : 2022-02-22
def get_platform_data(artist, platform, type, start_date, end_date, collect_item_list):
    '''
    <<input>>
    artist: 아티스트 이름
    platform: 플랫폼 이름
    type: 누적 or 기간별
    start_date: 시작날짜
    end_date: 끝날짜 (type=누적인 경우 null)
    collect_item_list: 해당 플랫폼,아티스트의 조사항목 list

    <<output>>
    크롤링 데이터 list
    '''
    # return list
    filter_datas = []
    # list 길이 초기화
    for i in collect_item_list:
        filter_datas.append("NULL")

    if type == "누적":
        start_date_dateobject = datetime.datetime.strptime(start_date, "%Y-%m-%d")
        start_date_string = start_date_dateobject.strftime("%Y-%m-%d")
        filter_objects = CollectData.objects.filter(collect_items__artist=artist, collect_items__platform=platform,
            collect_items__reserved_date = start_date_string)
        if filter_objects.exists():
            filter_value = filter_objects.values().first()
            # 아티스트-플랫폼가 해당날짜에 크롤링한 첫번째 데이터
            filter_value = filter_value["collect_items"]
            # 숫자 데이터, 날짜 데이터만 보내주기(id 등 제외)
            for i, field_name in enumerate(collect_item_list):
                # 수집항목에 포함되는 것만 list에 포함
                if not field_name in filter_value:
                    continue
                else:
                    # 숫자 데이터, 문자열로 된 숫자 데이터(예: "123")
                    if isinstance(filter_value[field_name], int) or filter_value[field_name].isdigit():
                        filter_datas[i] = int(filter_value[field_name])
                    # 날짜 데이터(0000-00-00형태로 반환)
                    else:
                        tmpdate = parse(filter_value[field_name])
                        filter_datas[i] = tmpdate.strftime("%Y-%m-%d")
            return filter_datas
        else:
            return filter_datas
    elif type == "기간별":
        start_date_dateobject = datetime.datetime.strptime(start_date, "%Y-%m-%d").date() - datetime.timedelta(1)
        end_date_dateobject = datetime.datetime.strptime(end_date, "%Y-%m-%d").date()
        start_date_string = start_date_dateobject.strftime("%Y-%m-%d")
        end_date_string = end_date_dateobject.strftime("%Y-%m-%d")
        filter_start_objects = CollectData.objects.filter(collect_items__artist=artist, collect_items__platform=platform,
            collect_items__reserved_date = start_date_string)
        filter_end_objects = CollectData.objects.filter(collect_items__artist=artist, collect_items__platform=platform,
            collect_items__reserved_date = end_date_string)
        if filter_start_objects.exists() and filter_end_objects.exists():
            filter_start_value = filter_start_objects.values().first()
            filter_end_value = filter_end_objects.values().first()
            # 아티스트-플랫폼가 시작날짜에 크롤링한 첫번째 데이터
            filter_start_value = filter_start_value["collect_items"]
            # 아티스트-플랫폼가 끝날짜에 크롤링한 첫번째 데이터
            filter_end_value = filter_end_value["collect_items"]
            # 숫자 데이터, 날짜 데이터만 보내주기(id 등 제외)
            for i, field_name in enumerate(collect_item_list):
                if "증감내역" in field_name:
                    continue
                # 끝날짜 데이터가 없다면 0으로
                if not field_name in filter_end_value:
                    filter_datas[i] = 0
                else:
                    # 숫자 데이터, 문자열로 된 숫자 데이터(예: "123")
                    if isinstance(filter_end_value[field_name], int) or filter_end_value[field_name].isdigit():
                        if field_name in filter_start_value:
                            # 뒤의 날짜 값
                            filter_datas[i] = int(filter_end_value[field_name])
                            if type=="기간별":
                                # 증감내역 데이터
                                filter_datas[i+1] = int(filter_end_value[field_name]) - int(filter_start_value[field_name])
                        # 시작날짜 데이터가 없는 경우
                        else: # 시작날짜값을 0으로 처리한 형태
                            filter_datas[i] = filter_end_value[field_name]
                            if type=="기간별":
                                # 증감내역 데이터
                                filter_datas[i+1] = 0
                    # 날짜 데이터(0000-00-00형태로 반환)
                    else:
                        tmpdate = parse(filter_end_value[field_name])
                        filter_datas[i] = tmpdate.strftime("%Y-%m-%d")
        return filter_datas
    else:
        return filter_datas

# 정의 : 데이터리포트 - 크롤링데이터 엑셀파일을 생성(추출)한다.
# 목적 : 크롤링데이터 엑셀 다운로드(export)를 위해 엑셀파일 생성
# 멤버함수 : get_platform_data
# 개발자 : 김민희, minheekim3@naver.com
# 최종수정일 : 2022-02-22
def export_datareport(excel_export_type, excel_export_start_date, excel_export_end_date):
    '''
    <<input>>
    excel_export_type: 누적 or 기간별
    excel_export_start_date: 시작날짜
    excel_export_end_date: 끝날짜 (type=누적인 경우 null)

    <<output>>
    엑셀 file
    '''
    # db에서 platform과 platform_target_item 가져오기
    db_platform_datas = []
    platforms = Platform.objects.all()
    if platforms.exists():
        platform_objects_values = platforms.values()
        for platform_value in platform_objects_values:
            collecttargets = CollectTarget.objects.filter(platform=platform_value["id"])
            collecttargets = collecttargets.values()
            collect_item = set()
            for collecttarget in collecttargets:
                platform_objects = CollectTargetItem.objects.filter(collect_target_id=collecttarget["id"])
                platform_objects_values = platform_objects.values()
                for p in platform_objects_values:
                    if p["target_name"] in collect_item:
                        continue
                    collect_item.add(p["target_name"])
                    if excel_export_type=="기간별" and p["target_name"] != 'user_created':
                        collect_item.add(p["target_name"]+" 의 증감내역")
            collect_item = list(collect_item)

            platform_header = []
            objects = CollectData.objects.filter(collect_items__platform=platform_value["name"])
            objects_values = objects.values()
            if len(objects_values) > 0:
                key_list = list(objects_values[0]["collect_items"].keys())
                for key in key_list:
                    if key in collect_item:
                        platform_header.append(key)
                        if excel_export_type=="기간별" and key != 'user_created':
                            platform_header.append(key+" 의 증감내역")
                    else:
                        continue
            else:
                platform_header = collect_item
            db_platform_datas.append({
                "platform": platform_value["name"],
                "collect_item": platform_header
            })

    # db에서 artist 가져오기
    db_artists = []
    artists = Artist.objects.all()
    if artists.exists():
        artist_objects_values = list(artists.values())
        artist_objects_values = sorted(sorted(artist_objects_values, key=lambda c:c['name']), key=lambda c:0 if re.search('[ㄱ-힣]', c['name'][0]) else 1)
        for artist_value in artist_objects_values:
            db_artists.append(artist_value["name"])

    # return할 excel file 생성
    book = openpyxl.Workbook()
    sheet = book.active

    row = 1
    col = 1

    thick_border = Border(left=Side(style="medium"),
                          right=Side(style="medium"),
                          top=Side(style="medium"),
                          bottom=Side(style="medium"))

    sheet.cell(row=1, column=1).value = "플랫폼"
    sheet.cell(row=1, column=1).border = thick_border
    sheet.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")
    sheet.cell(row=1, column=1).font = fonts.Font(bold=True)

    sheet.cell(row=2, column=1).value = "아티스트"
    sheet.cell(row=2, column=1).border = thick_border
    sheet.cell(row=2, column=1).alignment = Alignment(horizontal="center", vertical="center")
    sheet.cell(row=2, column=1).font = fonts.Font(bold=True)
    col += 1
    sheet.row_dimensions[2].height = 35
    sheet.column_dimensions["A"].width = 25

    # 플랫폼 이름과 수집항목 띄우기
    for data in db_platform_datas:
        if len(data["collect_item"]) == 0:
            continue
        sheet.cell(row=row, column=col).value = data["platform"]
        sheet.merge_cells(start_row=row, start_column=col,
                          end_row=row, end_column=col+len(data["collect_item"])-1)
        sheet.cell(row=row, column=col).border = thick_border
        sheet.cell(row=row, column=col).alignment = Alignment(horizontal="center", vertical="center")
        sheet.cell(row=row, column=col).font = fonts.Font(bold=True)

        for i, collect_data in enumerate(data["collect_item"]):
            sheet.column_dimensions[get_column_letter(col+i)].width = 20
            sheet.cell(row=row+1, column=col+i).value = collect_data
            sheet.cell(row=row+1, column=col+i).border = thick_border
            sheet.cell(row=row+1, column=col+i).alignment = Alignment(horizontal="center", vertical="center")
        col += len(data["collect_item"])

    # 아티스트별로 플랫폼순서대로 가져와서 띄우기
    row = 3
    col = 1
    for artist in db_artists:
        artist_name = artist
        sheet.cell(row=row, column=col).value = artist_name
        sheet.cell(row=row, column=col).border = Border(right=Side(style="medium"))
        col += 1
        for platform in db_platform_datas:
            # 아티스트의 플랫폼마다의 정보 가져오기
            platform_name = platform["platform"]
            platform_data_list = get_platform_data(artist=artist_name, platform=platform_name,
                                                   type=excel_export_type, start_date=excel_export_start_date, end_date=excel_export_end_date,
                                                   collect_item_list=platform["collect_item"])
            for i, platform_data in enumerate(platform_data_list):
                if platform_data is None or platform_data == "NULL":
                    # null이면 shade 처리
                    sheet.cell(row=row, column=col+i).fill = PatternFill(start_color="C4C4C4", end_color="C4C4C4", fill_type="solid")
                else:
                    sheet.cell(row=row, column=col+i).value = platform_data
            sheet.cell(row=row, column=col+len(platform["collect_item"])-1).border = Border(right=Side(style="medium"))
            col += len(platform["collect_item"])
        row += 1
        col = 1
    return book

# 정의 : 데이터리포트 - 크롤링데이터 엑셀파일로부터 db의 collect_data table에 저장한다.
# 목적 : 크롤링데이터 엑셀 업로드(import)를 통해 db에 크롤링데이터 저장
# 멤버함수 : save_collect_data_target
# 개발자 : 김민희, minheekim3@naver.com
# 최종수정일 : 2022-02-22
def import_datareport(worksheet, excel_import_date):
    '''
    <<input>>
    worksheet: 엑셀 파일 sheet
    excel_import_date: db에 저장하고자하는 날짜

    <<output>>
    x
    '''
    platform_data_list = []
    row_num = 0

    if excel_import_date is None:
        # 날짜가 없다면 default는 해당날짜
        target_date = datetime.datetime.today()
    else:
        target_date = datetime.datetime.strptime(excel_import_date, "%Y-%m-%d")
    target_date = datetime.date(target_date.year, target_date.month, target_date.day)
    target_date_string = target_date.strftime("%Y-%m-%d")

    # 엑셀파일 파싱
    for row in worksheet.iter_rows():
        # 플랫폼 정보 나열
        if row_num == 0:
            item_num = 0
            platform_name = "platform"
            data_json = {}
            for cell in row:
                platform_value = str(cell.value)
                if platform_value == "플랫폼":
                    platform_name = "플랫폼"
                elif platform_value != "None":
                    if platform_name != "플랫폼":
                        # 새로운 게 등장한 거므로 저장
                        data_json["platform"] = platform_name
                        data_json["item_num"] = item_num
                        data_json["item_list"] = []
                        platform_data_list.append(data_json)
                        data_json = {}
                    item_num = 1
                    platform_name = platform_value
                else:  # None일때
                    item_num += 1
            # platform_data_list에 저장
            data_json["platform"] = platform_name
            data_json["item_num"] = item_num
            data_json["item_list"] = []
            platform_data_list.append(data_json)
            row_num += 1
        # 아티스트 정보 나열
        elif row_num == 1:
            platform_index = 0
            current_index = 0
            for cell in row:
                collect_value = str(cell.value)
                if collect_value != "아티스트":
                    platform_data_list[platform_index]["item_list"].append(collect_value)
                    current_index += 1
                    if current_index >= platform_data_list[platform_index]["item_num"]:
                        platform_index += 1
                        current_index = 0
            row_num += 1
        # 크롤링데이터 정보 나열
        else:
            platform_index = 0
            current_index = 0
            artist_name = "artist"
            data_json = {}
            for i, cell in enumerate(row):
                # 아티스트 이름 나열
                if i == 0:
                    artist_name = str(cell.value)
                else:
                    # 아티스트와 플랫폼 이름에 대해 업데이트
                    platform_name = platform_data_list[platform_index]["platform"]
                    collect_value = platform_data_list[platform_index]["item_list"][current_index]
                    value = str(cell.value)
                    if value != "None":
                        # -가 있으면 날짜
                        if "-" in value:
                            value = str(cell.value)
                        # ,가 있으면 날짜
                        elif "," in value:
                            dateobject = parse(str(cell.value))
                            value = "%s-%s-%s"%(dateobject.year, dateobject.month, dateobject.day)
                        else:
                            value = int(cell.value)
                        data_json[collect_value] = value
                    current_index += 1
                    if current_index >= platform_data_list[platform_index]["item_num"]:
                        # 데이터 저장
                        data_json["artist"] = artist_name
                        save_collect_data_target(data_json, platform_name, target_date_string)
                        platform_index += 1
                        current_index = 0
                        data_json = {}

# 정의 : 데이터리포트 - 수집정보 엑셀파일로부터 db의 platform, artist, collect_target, collect_target_item, schedule table에 저장한다.
# 목적 : 수집정보 엑셀 업로드(import)를 통해 db에 수집정보 저장
# 멤버함수 : save_platform, save_artist, save_collect_target, save_schedule, save_collect_target_item
# 개발자 : 김민희, minheekim3@naver.com
# 최종수정일 : 2022-02-22
def import_collects(worksheet):
    '''
    <<input>>
    worksheet: 엑셀 파일 sheet

    <<output>>
    x
    '''
    platform_data_list = []
    artist_data_list = []
    collect_target_data_list = []
    collect_target_item_data_list = []
    platform_start_index = 7
    row_num = 0

    # 엑셀파일 파싱
    for row in worksheet.iter_rows():
        if row_num == 0 or row_num == 3 or row_num == 4:
            row_num += 1
            continue
        # 플랫폼 정보 나열
        if row_num == 1:
            item_num = 0
            platform_name = "플랫폼"
            data_json = {}
            for i, cell in enumerate(row):
                if i < platform_start_index:
                    continue
                platform_value = str(cell.value)
                if platform_value != "None":
                    if platform_name != "플랫폼":
                        # 새로운 게 등장한 거므로 저장
                        data_json["platform"] = platform_name
                        data_json["item_num"] = item_num
                        data_json["item_list"] = []
                        data_json["item_xpath_list"] = []
                        platform_data_list.append(data_json)
                        data_json = {}
                    item_num = 1
                    platform_name = platform_value
                else:  # None일때
                    item_num += 1
            # platform_data_list에 저장
            data_json["platform"] = platform_name
            data_json["item_num"] = item_num
            data_json["item_list"] = []
            data_json["item_xpath_list"] = []
            platform_data_list.append(data_json)
            row_num += 1
        # 플랫폼 주소 나열
        elif row_num == 2:
            platform_index = 0
            current_index = 0
            for i, cell in enumerate(row):
                if i < platform_start_index:
                    continue
                collect_value = str(cell.value)
                current_index += 1
                if collect_value != "None":
                    platform_data_list[platform_index]["url"] = collect_value
                if current_index >= platform_data_list[platform_index]["item_num"]:
                    platform_index += 1
                    current_index = 0
            row_num += 1
        # 지표 정보(영어) 나열
        elif row_num == 5:
            platform_index = 0
            current_index = 0
            for i, cell in enumerate(row):
                if i < platform_start_index:
                    continue
                collect_value = str(cell.value)
                if collect_value != "지표 이름":
                    platform_data_list[platform_index]["item_list"].append(collect_value)
                    current_index += 1
                    if current_index >= platform_data_list[platform_index]["item_num"]:
                        platform_index += 1
                        current_index = 0
            row_num += 1
        # 아티스트 & target_url 정보 나열
        else:
            platform_index = 0
            current_index = 0
            data_json = {}
            data_json2 = {}
            data_json3 = {}
            artist_name = ""
            for i, cell in enumerate(row):
                # 아티스트 이름 나열
                if i == 0:  # 이름
                    if str(cell.value) == "None":
                        break
                    data_json["name"] = str(cell.value)
                    artist_name = str(cell.value)
                elif i == 1:  # 구분
                    if str(cell.value) != "None":
                        data_json["level"] = str(cell.value)
                elif i == 2:  # 성별
                    if str(cell.value) != "None":
                        data_json["gender"] = str(cell.value)
                elif i == 3:  # 멤버수
                    if str(cell.value) != "None":
                        data_json["member_num"] = int(cell.value)
                elif i == 4:  # 국적
                    if str(cell.value) != "None":
                        data_json["member_nationality"] = str(cell.value)
                elif i == 5:  # 기획사
                    if str(cell.value) != "None":
                        data_json["agency"] = str(cell.value)
                elif i == 6:  # 데뷔일
                    if str(cell.value) != "None":
                        strings = str(cell.value).split(" ")
                        data_json["debut_date"] = strings[0]
                    artist_data_list.append(data_json)
                    data_json = {}
                # target_url나올 때
                else:
                    collect_value = str(cell.value)
                    current_index += 1
                    if "http" in collect_value:
                        if "target_url" in data_json2:
                            data_json2["target_url_2"] = collect_value
                        else:
                            data_json2["target_url"] = collect_value
                    elif collect_value != " " and collect_value != "None" and platform_data_list[platform_index]["item_list"][current_index-1] != "None":
                        data_json3 = {
                            "platform": platform_data_list[platform_index]["platform"],
                            "artist": artist_name,
                            "target_name": platform_data_list[platform_index]["item_list"][current_index-1],
                            "xpath": collect_value
                        }
                        collect_target_item_data_list.append(data_json3)
                    if current_index >= platform_data_list[platform_index]["item_num"]:
                        platform_index += 1
                        current_index = 0
                        collect_target_data_list.append(data_json2)
                        data_json2 = {}

    # db에 저장
    # platform table에 저장
    for platform_data in platform_data_list:
        save_platform({
            "name": platform_data["platform"],
            "url": platform_data["url"]
        })
    # artist table에 저장
    collect_target_index = 0
    for artist_data in artist_data_list:
        save_artist(artist_data)
        # collect_target 저장
        for platform_data in platform_data_list:
            if "target_url" in collect_target_data_list[collect_target_index]:
                platform_filter_object = Platform.objects.filter(name=platform_data["platform"])
                platform_filter_object = platform_filter_object.values().first()
                artist_filter_object = Artist.objects.filter(name=artist_data["name"])
                artist_filter_object = artist_filter_object.values().first()
                if 'target_url_2' in collect_target_data_list[collect_target_index]:
                    save_collect_target({
                        "artist": artist_filter_object['id'],
                        "platform": platform_filter_object['id'],
                        "target_url": collect_target_data_list[collect_target_index]['target_url'],
                        "target_url_2": collect_target_data_list[collect_target_index]['target_url_2'],
                    })
                else:
                    save_collect_target({
                        "artist": artist_filter_object["id"],
                        "platform": platform_filter_object["id"],
                        "target_url": collect_target_data_list[collect_target_index]["target_url"]
                    })
                collecttarget_object = CollectTarget.objects.filter(artist_id = artist_filter_object['id'],
                    platform=platform_filter_object['id'])
                collecttarget_object = collecttarget_object.values().first()
                collecttargetid = collecttarget_object["id"]
                # schedule table에 저장
                save_schedule(collecttargetid)
            collect_target_index += 1
    # collect_target_item table에 저장
    for collect_target_item_data in collect_target_item_data_list:
        save_collect_target_item(collect_target_item_data)


# 정의 : 데이터리포트 - 로그인정보 엑셀파일로부터 db의 auth_info table에 저장한다.
# 목적 : 로그인정보 엑셀 업로드(import)를 통해 db에 수집정보 저장
# 멤버함수 : save_auth_info
# 개발자 : 김민희, minheekim3@naver.com
# 최종수정일 : 2022-02-22
def import_authinfo(worksheet):
    '''
    <<input>>
    worksheet: 엑셀 파일 sheet

    <<output>>
    x
    '''
    platform_data_list = []
    artist_data_list = []
    auth_info_list = []
    platform_start_index = 1
    row_num = 0

    # 엑셀파일 파싱
    for row in worksheet.iter_rows():
        if row_num == 0 or row_num == 3 or row_num == 4:
            row_num += 1
            continue
        # 플랫폼 정보 나열
        if row_num == 1:
            item_num = 0
            platform_name = "플랫폼"
            data_json = {}
            for i, cell in enumerate(row):
                if i < platform_start_index:
                    continue
                platform_value = str(cell.value)
                if platform_value != "None":
                    if platform_name != "플랫폼":
                        # 새로운 게 등장한 거므로 저장
                        data_json["platform"] = platform_name
                        data_json["item_num"] = item_num
                        data_json["item_list"] = []
                        platform_data_list.append(data_json)
                        data_json = {}
                    item_num = 1
                    platform_name = platform_value
                else:  # None일때
                    item_num += 1
            # platform_data_list에 저장
            data_json["platform"] = platform_name
            data_json["item_num"] = item_num
            data_json["item_list"] = []
            platform_data_list.append(data_json)
            row_num += 1
        # 플랫폼 주소 나열
        elif row_num == 2:
            platform_index = 0
            current_index = 0
            for i, cell in enumerate(row):
                if i < platform_start_index:
                    continue
                collect_value = str(cell.value)
                current_index += 1
                if collect_value != "None":
                    platform_data_list[platform_index]["url"] = collect_value
                if current_index >= platform_data_list[platform_index]["item_num"]:
                    platform_index += 1
                    current_index = 0
            row_num += 1
        # 지표 정보(영어) 나열
        elif row_num == 5:
            platform_index = 0
            current_index = 0
            for i, cell in enumerate(row):
                if i < platform_start_index:
                    continue
                collect_value = str(cell.value)
                if collect_value != "지표 이름":
                    platform_data_list[platform_index]["item_list"].append(collect_value)
                    current_index += 1
                    if current_index >= platform_data_list[platform_index]["item_num"]:
                        platform_index += 1
                        current_index = 0
            row_num += 1
        # 아티스트 & auth_info 정보 나열
        else:
            platform_index = 0
            current_index = 0
            data_json = {}
            data_json2 = {}
            artist_name = ""
            for i, cell in enumerate(row):
                # 아티스트 이름 나열
                if i == 0:  # 이름
                    if str(cell.value) == "None":
                        break
                    data_json["name"] = str(cell.value)
                    artist_name = str(cell.value)
                    artist_data_list.append(artist_name)
                else: #auth_info
                    collect_value = str(cell.value)
                    if collect_value != 'null' and collect_value != 'None' and collect_value != ' ':
                        data_json2[platform_data_list[platform_index]["item_list"][current_index]] = collect_value
                    current_index += 1
                    if current_index >= platform_data_list[platform_index]["item_num"]:
                        if data_json2 != {}:
                            data_json2["artist"] = artist_name
                            data_json2["platform"] = platform_data_list[platform_index]["platform"]
                            auth_info_list.append(data_json2)
                            data_json2 = {}
                        platform_index += 1
                        current_index = 0

    # db에 저장하는 부분
    # auth_info table에 저장
    for auth_info_item in auth_info_list:
        save_auth_info(auth_info_item)


# 정의 : db의 auth_info table에 데이터를 저장한다.
# 목적 : import_authinfo 함수 내에서 사용
# 멤버함수 : 
# 개발자 : 김민희, minheekim3@naver.com
# 최종수정일 : 2022-02-22
def save_auth_info(data_json):
    '''
    기능: 로그인정보 저장

    <<input>>
    data_json: auth_info에 저장하고자하는 json형태의 데이터

    <<output>>
    x
    '''
    platform = data_json["platform"]
    artist = data_json["artist"]
    artist_object = Artist.objects.filter(name = artist).first()
    platform_object = Platform.objects.filter(name = platform).first()
    collect_target_object = CollectTarget.objects.filter(platform_id = platform_object.id, artist_id = artist_object.id).first()
    new_data = {'collect_target': collect_target_object.id}
    for key in data_json.keys():
        if key == 'platform' or key == 'artist':
            continue
        new_data[key] = data_json[key]
    obj = AuthInfo.objects.filter(collect_target=collect_target_object.id).first()
    authinfo_serializer = AuthInfoSerializer(obj, data=new_data)
    if authinfo_serializer.is_valid():
        authinfo_serializer.save()


# 정의 : db의 collect_data table에 데이터를 저장한다.
# 목적 : import_datareport 함수 내에서 사용
# 멤버함수 : 
# 개발자 : 김민희, minheekim3@naver.com
# 최종수정일 : 2022-02-22
def save_collect_data_target(data_json, platform, target_date_string):
    '''
    기능: CollectData table에 생성 및 업데이트

    <<input>>
    data_json: collect_data에 저장하고자하는 json형태의 크롤링데이터
    platform: 플랫폼 이름
    target_date_string: 수집날짜

    <<output>>
    x
    '''
    # 조사항목이 없으므로 저장하지 않음
    if len(data_json.keys())==1:
        return
    # collectdata 저장
    data_json["reserved_date"] = target_date_string
    data_json["platform"] = platform
    artist_object = Artist.objects.filter(name=data_json["artist"])
    artist_object = artist_object.values()[0]
    platform_object = Platform.objects.filter(name=platform)
    platform_object = platform_object.values()[0]
    collecttarget_object = CollectTarget.objects.filter(platform_id = platform_object["id"], artist_id = artist_object["id"])
    collecttarget_object = collecttarget_object.values()[0]
    # 해당 날짜의 데이터 생성 및 업데이트
    CollectData.objects.update_or_create(
        collect_target_id = collecttarget_object['id'],
        collect_items__reserved_date = target_date_string,
        # 업데이트하고자하는 값
        defaults = {"collect_items": data_json}
    )


# 정의 : db의 platform table에 데이터를 저장한다.
# 목적 : import_collects 함수 내에서 사용
# 멤버함수 : 
# 개발자 : 김민희, minheekim3@naver.com
# 최종수정일 : 2022-02-22
def save_platform(data_json):
    '''
    기능: platform table에 생성 및 업데이트

    <<input>>
    data_json: platform에 저장하고자하는 json형태의 크롤링데이터

    <<output>>
    x
    '''
    obj = Platform.objects.filter(name=data_json["name"]).first()
    # 생성 및 업데이트
    platform_serializer = PlatformSerializer(obj, data=data_json)
    if platform_serializer.is_valid():
        platform_serializer.save()


# 정의 : db의 artist table에 데이터를 저장한다.
# 목적 : import_collects 함수 내에서 사용
# 멤버함수 : 
# 개발자 : 김민희, minheekim3@naver.com
# 최종수정일 : 2022-02-22
def save_artist(data_json):
    '''
    기능: artist table에 생성 및 업데이트

    <<input>>
    data_json: artist에 저장하고자하는 json형태의 크롤링데이터

    <<output>>
    x
    '''
    obj = Artist.objects.filter(name=data_json["name"]).first()
    # 생성 및 업데이트
    artist_serializer = ArtistSerializer(obj, data=data_json)
    if artist_serializer.is_valid():
        artist_serializer.save()


# 정의 : db의 collect_target_item table에 데이터를 저장한다.
# 목적 : import_collects 함수 내에서 사용
# 멤버함수 : 
# 개발자 : 김민희, minheekim3@naver.com
# 최종수정일 : 2022-02-22
def save_collect_target_item(data_json):
    '''
    기능: CollectTargetItem(조사항목) table에 생성 및 업데이트

    <<input>>
    data_json: collect_target_item에 저장하고자하는 json형태의 크롤링데이터

    <<output>>
    x
    '''
    artist_object = Artist.objects.filter(name=data_json["artist"])
    artist_object = artist_object.values()[0]
    platform_object = Platform.objects.filter(name=data_json["platform"])
    platform_object = platform_object.values()[0]
    collecttarget_object = CollectTarget.objects.filter(artist_id=artist_object["id"], platform_id=platform_object["id"])
    collecttarget_object = collecttarget_object.values()[0]
    obj = CollectTargetItem.objects.filter(collect_target_id=collecttarget_object["id"], target_name=data_json["target_name"]).first()
    data_json = {
        "collect_target": collecttarget_object["id"],
        "target_name": data_json["target_name"],
        "xpath": data_json["xpath"]
    }
    # 생성 및 업데이트
    target_item_serializer = CollectTargetItemSerializer(obj, data=data_json)
    if target_item_serializer.is_valid():
        target_item_serializer.save()
    

# 정의 : db의 collect_target table에 데이터를 저장한다.
# 목적 : import_collects 함수 내에서 사용
# 멤버함수 : 
# 개발자 : 김민희, minheekim3@naver.com
# 최종수정일 : 2022-02-22
def save_collect_target(data_json):
    '''
    기능: CollectTarget(조사대상) table에 생성 및 업데이트

    <<input>>
    data_json: collect_target에 저장하고자하는 json형태의 크롤링데이터

    <<output>>
    x
    '''
    obj = CollectTarget.objects.filter(platform=data_json["platform"], artist=data_json["artist"]).first()
    # 생성 및 업데이트
    collect_target_item_serializer = CollectTargetSerializer(obj, data=data_json)
    if collect_target_item_serializer.is_valid():
        collect_target_item_serializer.save()


# 정의 : db의 schedule table에 데이터를 저장한다.
# 목적 : import_collects 함수 내에서 사용
# 멤버함수 : 
# 개발자 : 김민희, minheekim3@naver.com
# 최종수정일 : 2022-02-22
def save_schedule(collecttargetid):
    '''
    기능: Schedule(조사대상) table에 생성 및 업데이트

    <<input>>
    collecttargetid: 만들고자 하는 schedule이 참조하는 collect_target_id 값

    <<output>>
    x
    '''
    schedule_object = Schedule.objects.filter(collect_target_id = collecttargetid).first()
    schedule_data = {
        "collect_target": collecttargetid,
        "schedule_type": "daily",
        "active": True
    }
    # 생성 및 업데이트
    schedule_serializer = ScheduleSerializer(schedule_object, data=schedule_data)
    if schedule_serializer.is_valid():
        schedule_serializer.save()
